#!/usr/bin/env python3
"""
Container Security Scanner and Report Generator

This script scans Docker containers for vulnerabilities using Grype and generates
comprehensive reports in multiple formats for various use cases.

Features:
- Scans running Docker containers
- Categorizes images (Chainguard vs Legacy)
- Generates CSV files (for data processing)
- Generates HTML reports (for web viewing, Instruqt-compatible)
- Generates text reports (for terminal viewing)
- Generates formatted Excel reports with charts and color-coding
- Creates comparison reports
"""

import argparse
import json
import subprocess
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import PieChart, Reference
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class VulnerabilityScanner:
    """Main scanner class that orchestrates scanning and reporting."""
    
    # Images to categorize as Chainguard
    CHAINGUARD_PATTERNS = [
        "three-tier-nginx-cg",
        "three-tier-frontend-cg",
        "three-tier-backend-cg",
        "three-tier-db-cg",
        "cgr.dev/chainguard"
    ]
    
    SEVERITY_ORDER = ['Critical', 'High', 'Medium', 'Low', 'Negligible', 'Unknown']
    SEVERITY_COLORS = {
        'Critical': 'DC143C',
        'High': 'FF6347',
        'Medium': 'FFA500',
        'Low': 'FFD700',
        'Negligible': 'D3D3D3',
        'Unknown': 'C0C0C0'
    }
    
    def __init__(self, output_dir: str = "./scanners/scan-results"):
        """Initialize scanner with output directory."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.chainguard_results = []
        self.legacy_results = []
        self.scanned_images = []  # Track all scanned images (even those with 0 vulns)
        self.scan_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    def check_dependencies(self) -> bool:
        """Check if required tools are available."""
        dependencies = {
            'docker': 'Docker is required for container inspection',
            'grype': 'Grype is required for vulnerability scanning (https://github.com/anchore/grype)'
        }
        
        missing = []
        for cmd, description in dependencies.items():
            if not self._command_exists(cmd):
                missing.append(f"  ❌ {cmd}: {description}")
        
        if missing:
            print("Missing dependencies:")
            print("\n".join(missing))
            return False
        return True
    
    def _command_exists(self, command: str) -> bool:
        """Check if a command exists in PATH."""
        try:
            # Use 'where' on Windows, 'which' on Unix
            check_cmd = 'where' if sys.platform.startswith('win') else 'which'
            subprocess.run(
                [check_cmd, command],
                capture_output=True,
                check=True,
                encoding='utf-8',
                errors='replace'
            )
            return True
        except subprocess.CalledProcessError:
            return False
    
    def get_running_containers(self) -> List[Tuple[str, str]]:
        """Get list of running containers with their image names."""
        print("🔍 Detecting running containers...")
        try:
            result = subprocess.run(
                ['docker', 'compose', 'ps', '-q'],
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                check=True
            )
            container_ids = [cid.strip() for cid in result.stdout.split('\n') if cid.strip()]
            
            if not container_ids:
                print("⚠️  No running containers found via docker compose.")
                return []
            
            print(f"   Found {len(container_ids)} running container(s)")
            
            # Get image names and display them
            containers_with_images = []
            for container_id in container_ids:
                image_name = self.get_image_name(container_id)
                if image_name:
                    containers_with_images.append((container_id, image_name))
                    platform = self.get_image_platform(image_name)
                    platform_info = f" [{platform}]" if platform else ""
                    print(f"      • {image_name}{platform_info}")
            
            return containers_with_images
            
        except subprocess.CalledProcessError as e:
            print(f"❌ Error getting containers: {e}")
            return []
    
    def get_image_name(self, container_id: str) -> Optional[str]:
        """Get image name for a container."""
        try:
            result = subprocess.run(
                ['docker', 'inspect', '--format={{.Config.Image}}', container_id],
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                check=True
            )
            return result.stdout.strip()
        except subprocess.CalledProcessError:
            return None
    
    def get_image_platform(self, image_name: str) -> Optional[str]:
        """Get the platform/architecture of an image."""
        try:
            result = subprocess.run(
                ['docker', 'inspect', '--format={{.Os}}/{{.Architecture}}', image_name],
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                check=True
            )
            return result.stdout.strip()
        except subprocess.CalledProcessError:
            return None
    
    def is_chainguard_image(self, image_name: str) -> bool:
        """Check if image matches Chainguard patterns."""
        return any(pattern in image_name for pattern in self.CHAINGUARD_PATTERNS)
    
    def scan_image(self, image_name: str) -> List[Dict]:
        """Scan an image with Grype and return vulnerabilities."""
        platform = self.get_image_platform(image_name)
        platform_info = f" [{platform}]" if platform else ""
        print(f"   Scanning: {image_name}{platform_info}")
        
        try:
            result = subprocess.run(
                ['grype', image_name, '-o', 'json'],
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                check=True
            )
            
            data = json.loads(result.stdout)
            vulnerabilities = []
            
            # Parse matches
            for match in data.get('matches', []):
                vuln = {
                    'Image': image_name,
                    'Platform': platform or 'unknown',
                    'Package': match['artifact']['name'],
                    'Version': match['artifact']['version'],
                    'Vulnerability': match['vulnerability']['id'],
                    'Severity': match['vulnerability']['severity'],
                    'Type': match['artifact']['type'],
                    'FixedInVersion': match['vulnerability'].get('fixedInVersion', 'N/A')
                }
                vulnerabilities.append(vuln)
            
            return vulnerabilities
            
        except subprocess.CalledProcessError as e:
            print(f"      ⚠️  Error scanning {image_name}: {e}")
            return []
        except json.JSONDecodeError as e:
            print(f"      ⚠️  Error parsing Grype output for {image_name}: {e}")
            return []
    
    def scan_all_containers(self):
        """Scan all running containers and categorize results."""
        containers = self.get_running_containers()
        if not containers:
            return
        
        print("\n🔬 Scanning containers for vulnerabilities...")
        print("=" * 60)
        
        seen_images = set()
        
        for container_id, image_name in containers:
            # Skip duplicates
            if image_name in seen_images:
                print(f"   ⏭️  Skipping already scanned: {image_name}")
                continue
            
            seen_images.add(image_name)
            
            # Track this image (even if it has 0 vulnerabilities)
            self.scanned_images.append(image_name)
            
            # Scan the image
            vulnerabilities = self.scan_image(image_name)
            
            if vulnerabilities:
                # Categorize as Chainguard or Legacy
                if self.is_chainguard_image(image_name):
                    self.chainguard_results.extend(vulnerabilities)
                    print(f"      📦 Found {len(vulnerabilities)} vulnerabilities (Chainguard)")
                else:
                    self.legacy_results.extend(vulnerabilities)
                    print(f"      📦 Found {len(vulnerabilities)} vulnerabilities (Legacy)")
            else:
                print(f"      ✅ No vulnerabilities found")
        
        print("=" * 60)
    
    def write_csv(self, results: List[Dict], filename: str) -> Optional[Path]:
        """Write results to CSV file."""
        if not results:
            return None
        
        filepath = self.output_dir / filename
        
        # CSV headers
        headers = ['Image', 'Platform', 'Package', 'Version', 'Vulnerability', 'Severity', 'Type', 'FixedInVersion']
        
        with open(filepath, 'w') as f:
            # Write header
            f.write(','.join(headers) + '\n')
            
            # Write data
            for row in results:
                line = ','.join([f'"{row[h]}"' for h in headers])
                f.write(line + '\n')
        
        return filepath
    
    def save_csv_reports(self):
        """Save CSV reports for both categories."""
        print("\n💾 Saving CSV reports...")
        
        # Check which image types were actually scanned
        has_chainguard_images = any(self.is_chainguard_image(img) for img in self.scanned_images)
        has_legacy_images = any(not self.is_chainguard_image(img) for img in self.scanned_images)
        
        # Only report on categories that were actually scanned
        if has_chainguard_images:
            if self.chainguard_results:
                filepath = self.write_csv(self.chainguard_results, 'grype-chainguard-images.csv')
                print(f"   ✅ Chainguard: {filepath} ({len(self.chainguard_results)} vulnerabilities)")
            else:
                print("   ✅ Chainguard: No vulnerabilities found! 🎉")
        
        if has_legacy_images:
            if self.legacy_results:
                filepath = self.write_csv(self.legacy_results, 'grype-legacy-images.csv')
                print(f"   ✅ Legacy: {filepath} ({len(self.legacy_results)} vulnerabilities)")
            else:
                print("   ✅ Legacy: No vulnerabilities found!")
    
    def write_html_report(self, results: List[Dict], filename: str, report_type: str = "") -> Optional[Path]:
        """Generate HTML report with styling and interactive elements."""
        filepath = self.output_dir / filename
        
        # Calculate statistics
        stats = {
            'total': len(results),
            'by_severity': Counter(d['Severity'] for d in results),
            'by_image': Counter(d['Image'] for d in results),
            'by_type': Counter(d['Type'] for d in results),
            'unique_cves': len(set(d['Vulnerability'] for d in results)),
            'unique_packages': len(set(d['Package'] for d in results)),
            'unique_images': len(set(d['Image'] for d in results)),
            'fixable': sum(1 for d in results if d.get('FixedInVersion', 'N/A') != 'N/A'),
        }
        
        # Get all images (including those with 0 vulnerabilities)
        all_images_with_counts = {}
        for image in self.scanned_images:
            all_images_with_counts[image] = stats['by_image'].get(image, 0)
        
        # HTML template
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_type} Security Scan Report</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f7fa;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        
        h1 {{
            color: #1F4E78;
            margin-bottom: 10px;
            font-size: 2em;
        }}
        
        .timestamp {{
            color: #666;
            font-style: italic;
            margin-bottom: 30px;
            font-size: 0.9em;
        }}
        
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}
        
        .metric-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        .metric-card h3 {{
            font-size: 0.9em;
            opacity: 0.9;
            margin-bottom: 10px;
            font-weight: normal;
        }}
        
        .metric-card .value {{
            font-size: 2em;
            font-weight: bold;
        }}
        
        h2 {{
            color: #1F4E78;
            margin: 40px 0 20px 0;
            padding-bottom: 10px;
            border-bottom: 2px solid #e0e0e0;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        
        thead {{
            background: #1F4E78;
            color: white;
        }}
        
        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #e0e0e0;
        }}
        
        tbody tr:hover {{
            background: #f8f9fa;
        }}
        
        .severity {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 4px;
            font-weight: 600;
            font-size: 0.85em;
            text-align: center;
            min-width: 80px;
        }}
        
        .severity-Critical {{
            background: #DC143C;
            color: white;
        }}
        
        .severity-High {{
            background: #FF6347;
            color: white;
        }}
        
        .severity-Medium {{
            background: #FFA500;
            color: white;
        }}
        
        .severity-Low {{
            background: #FFD700;
            color: #333;
        }}
        
        .severity-Negligible {{
            background: #D3D3D3;
            color: #333;
        }}
        
        .severity-Unknown {{
            background: #C0C0C0;
            color: #333;
        }}
        
        .severity-summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        
        .severity-box {{
            padding: 15px;
            border-radius: 6px;
            text-align: center;
            color: white;
            font-weight: 600;
        }}
        
        .severity-box .count {{
            font-size: 2em;
            display: block;
            margin-bottom: 5px;
        }}
        
        .severity-box .label {{
            font-size: 0.9em;
            opacity: 0.9;
        }}
        
        a {{
            color: #0563C1;
            text-decoration: none;
        }}
        
        a:hover {{
            text-decoration: underline;
        }}
        
        .filter-bar {{
            margin: 20px 0;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 6px;
        }}
        
        .filter-bar input {{
            padding: 8px 12px;
            border: 1px solid #ddd;
            border-radius: 4px;
            width: 300px;
            font-size: 1em;
        }}
        
        .image-list {{
            list-style: none;
        }}
        
        .image-list li {{
            padding: 10px;
            margin: 5px 0;
            background: #f8f9fa;
            border-radius: 4px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        
        .image-list li:hover {{
            background: #e9ecef;
        }}
        
        .badge {{
            background: #667eea;
            color: white;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: 600;
        }}
        
        .badge.zero {{
            background: #28a745;
        }}
        
        .no-vulnerabilities {{
            text-align: center;
            padding: 60px 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 8px;
            margin: 20px 0;
        }}
        
        .no-vulnerabilities h2 {{
            color: white;
            border: none;
            font-size: 2em;
            margin: 0 0 15px 0;
        }}
        
        .no-vulnerabilities p {{
            font-size: 1.2em;
            opacity: 0.9;
        }}
        
        @media print {{
            body {{
                background: white;
            }}
            
            .container {{
                box-shadow: none;
            }}
            
            .filter-bar {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{report_type} Security Vulnerability Scan Report</h1>
        <div class="timestamp">Generated: {self.scan_timestamp}</div>
"""
        
        if not results:
            html += """
        <div class="no-vulnerabilities">
            <h2>✅ No Vulnerabilities Found!</h2>
            <p>All scanned images are clean and secure.</p>
        </div>
"""
        else:
            # Key metrics
            html += f"""
        <h2>Key Metrics</h2>
        <div class="metrics-grid">
            <div class="metric-card">
                <h3>Total Vulnerabilities</h3>
                <div class="value">{stats['total']}</div>
            </div>
            <div class="metric-card">
                <h3>Unique CVEs</h3>
                <div class="value">{stats['unique_cves']}</div>
            </div>
            <div class="metric-card">
                <h3>Affected Packages</h3>
                <div class="value">{stats['unique_packages']}</div>
            </div>
            <div class="metric-card">
                <h3>Affected Images</h3>
                <div class="value">{stats['unique_images']}</div>
            </div>
            <div class="metric-card">
                <h3>Fixable Issues</h3>
                <div class="value">{stats['fixable']}</div>
                <div style="font-size: 0.8em; margin-top: 5px;">
                    {stats['fixable']/stats['total']*100:.1f}% of total
                </div>
            </div>
        </div>
        
        <h2>Vulnerabilities by Severity</h2>
        <div class="severity-summary">
"""
            
            for severity in self.SEVERITY_ORDER:
                count = stats['by_severity'].get(severity, 0)
                if count > 0 or severity in ['Critical', 'High', 'Medium', 'Low']:
                    color = self.SEVERITY_COLORS.get(severity, 'CCCCCC')
                    html += f"""
            <div class="severity-box" style="background: #{color};">
                <span class="count">{count}</span>
                <span class="label">{severity}</span>
            </div>
"""
            
            html += """
        </div>
        
        <h2>Vulnerabilities by Image</h2>
        <ul class="image-list">
"""
            
            for image, count in sorted(all_images_with_counts.items(), key=lambda x: (-x[1], x[0])):
                platform = self.get_image_platform(image)
                platform_info = f" <span style='color: #666; font-size: 0.85em;'>[{platform}]</span>" if platform else ""
                badge_class = "zero" if count == 0 else ""
                html += f"""
            <li>
                <span>{image}{platform_info}</span>
                <span class="badge {badge_class}">{count} vulnerabilities</span>
            </li>
"""
            
            html += """
        </ul>
        
        <h2>Detailed Vulnerability Report</h2>
        <div class="filter-bar">
            <input type="text" id="searchInput" placeholder="🔍 Search vulnerabilities..." 
                   onkeyup="filterTable()">
        </div>
        
        <table id="vulnTable">
            <thead>
                <tr>
                    <th>Image</th>
                    <th>Platform</th>
                    <th>Package</th>
                    <th>Version</th>
                    <th>Vulnerability</th>
                    <th>Severity</th>
                    <th>Type</th>
                    <th>Fixed In</th>
                </tr>
            </thead>
            <tbody>
"""
            
            # Sort by severity
            severity_map = {sev: idx for idx, sev in enumerate(self.SEVERITY_ORDER)}
            sorted_results = sorted(results, key=lambda x: (severity_map.get(x['Severity'], 999), x['Image']))
            
            for vuln in sorted_results:
                cve = vuln['Vulnerability']
                if cve.startswith('CVE-'):
                    cve_link = f'<a href="https://nvd.nist.gov/vuln/detail/{cve}" target="_blank">{cve}</a>'
                elif cve.startswith('GHSA-'):
                    cve_link = f'<a href="https://github.com/advisories/{cve}" target="_blank">{cve}</a>'
                else:
                    cve_link = cve
                
                html += f"""
                <tr>
                    <td>{vuln['Image']}</td>
                    <td>{vuln.get('Platform', 'unknown')}</td>
                    <td>{vuln['Package']}</td>
                    <td>{vuln['Version']}</td>
                    <td>{cve_link}</td>
                    <td><span class="severity severity-{vuln['Severity']}">{vuln['Severity']}</span></td>
                    <td>{vuln['Type']}</td>
                    <td>{vuln.get('FixedInVersion', 'N/A')}</td>
                </tr>
"""
            
            html += """
            </tbody>
        </table>
"""
        
        # All scanned images section (even if no vulns)
        html += """
        <h2>All Scanned Images</h2>
        <ul class="image-list">
"""
        
        for image in sorted(self.scanned_images):
            count = all_images_with_counts.get(image, 0)
            platform = self.get_image_platform(image)
            platform_info = f" <span style='color: #666; font-size: 0.85em;'>[{platform}]</span>" if platform else ""
            badge_class = "zero" if count == 0 else ""
            html += f"""
            <li>
                <span>{image}{platform_info}</span>
                <span class="badge {badge_class}">{count} vulnerabilities</span>
            </li>
"""
        
        html += """
        </ul>
    </div>
    
    <script>
        function filterTable() {
            const input = document.getElementById('searchInput');
            const filter = input.value.toUpperCase();
            const table = document.getElementById('vulnTable');
            const tr = table.getElementsByTagName('tr');
            
            for (let i = 1; i < tr.length; i++) {
                let txtValue = tr[i].textContent || tr[i].innerText;
                if (txtValue.toUpperCase().indexOf(filter) > -1) {
                    tr[i].style.display = '';
                } else {
                    tr[i].style.display = 'none';
                }
            }
        }
    </script>
</body>
</html>
"""
        
        with open(filepath, 'w') as f:
            f.write(html)
        
        return filepath
    
    def write_text_report(self, results: List[Dict], filename: str, report_type: str = "") -> Optional[Path]:
        """Generate text report for terminal viewing."""
        filepath = self.output_dir / filename
        
        # Calculate statistics
        stats = {
            'total': len(results),
            'by_severity': Counter(d['Severity'] for d in results),
            'by_image': Counter(d['Image'] for d in results),
            'by_type': Counter(d['Type'] for d in results),
            'unique_cves': len(set(d['Vulnerability'] for d in results)),
            'unique_packages': len(set(d['Package'] for d in results)),
            'unique_images': len(set(d['Image'] for d in results)),
            'fixable': sum(1 for d in results if d.get('FixedInVersion', 'N/A') != 'N/A'),
        }
        
        # Get all images (including those with 0 vulnerabilities)
        all_images_with_counts = {}
        for image in self.scanned_images:
            all_images_with_counts[image] = stats['by_image'].get(image, 0)
        
        with open(filepath, 'w') as f:
            # Header
            f.write("=" * 80 + "\n")
            f.write(f"{report_type} SECURITY VULNERABILITY SCAN REPORT\n".center(80))
            f.write("=" * 80 + "\n")
            f.write(f"Generated: {self.scan_timestamp}\n".center(80))
            f.write("=" * 80 + "\n\n")
            
            if not results:
                f.write("\n")
                f.write("✅ NO VULNERABILITIES FOUND! ✅\n".center(80))
                f.write("\n")
                f.write("All scanned images are clean and secure.\n".center(80))
                f.write("\n")
            else:
                # Key Metrics
                f.write("KEY METRICS\n")
                f.write("-" * 80 + "\n")
                f.write(f"  Total Vulnerabilities:  {stats['total']}\n")
                f.write(f"  Unique CVEs:            {stats['unique_cves']}\n")
                f.write(f"  Affected Packages:      {stats['unique_packages']}\n")
                f.write(f"  Affected Images:        {stats['unique_images']}\n")
                if stats['total'] > 0:
                    f.write(f"  Fixable Issues:         {stats['fixable']} ({stats['fixable']/stats['total']*100:.1f}%)\n")
                else:
                    f.write(f"  Fixable Issues:         0 (0%)\n")
                f.write("\n")
                
                # Severity Breakdown
                f.write("VULNERABILITIES BY SEVERITY\n")
                f.write("-" * 80 + "\n")
                for severity in self.SEVERITY_ORDER:
                    count = stats['by_severity'].get(severity, 0)
                    if count > 0 or severity in ['Critical', 'High', 'Medium', 'Low']:
                        icon = "🔴" if severity == "Critical" else "🟠" if severity == "High" else "🟡" if severity == "Medium" else "🟢" if severity == "Low" else "⚪"
                        f.write(f"  {icon} {severity:12s}: {count:5d}\n")
                f.write("\n")
                
                # Image Breakdown
                f.write("VULNERABILITIES BY IMAGE\n")
                f.write("-" * 80 + "\n")
                for image, count in sorted(all_images_with_counts.items(), key=lambda x: (-x[1], x[0])):
                    platform = self.get_image_platform(image)
                    platform_info = f" [{platform}]" if platform else ""
                    f.write(f"  {image}{platform_info}\n")
                    f.write(f"    └─ {count} vulnerabilities\n")
                f.write("\n")
                
                # Package Type Breakdown
                f.write("VULNERABILITIES BY PACKAGE TYPE\n")
                f.write("-" * 80 + "\n")
                for pkg_type, count in sorted(stats['by_type'].items(), key=lambda x: x[1], reverse=True):
                    f.write(f"  {pkg_type:20s}: {count:5d}\n")
                f.write("\n")
                
                # Detailed Vulnerability List
                f.write("DETAILED VULNERABILITY REPORT\n")
                f.write("=" * 80 + "\n\n")
                
                # Sort by severity
                severity_map = {sev: idx for idx, sev in enumerate(self.SEVERITY_ORDER)}
                sorted_results = sorted(results, key=lambda x: (severity_map.get(x['Severity'], 999), x['Image']))
                
                for vuln in sorted_results:
                    f.write(f"Image:          {vuln['Image']}\n")
                    f.write(f"Platform:       {vuln.get('Platform', 'unknown')}\n")
                    f.write(f"Package:        {vuln['Package']} ({vuln['Version']})\n")
                    f.write(f"Vulnerability:  {vuln['Vulnerability']}\n")
                    
                    # Add URL for CVE/GHSA
                    if vuln['Vulnerability'].startswith('CVE-'):
                        f.write(f"URL:            https://nvd.nist.gov/vuln/detail/{vuln['Vulnerability']}\n")
                    elif vuln['Vulnerability'].startswith('GHSA-'):
                        f.write(f"URL:            https://github.com/advisories/{vuln['Vulnerability']}\n")
                    
                    f.write(f"Severity:       {vuln['Severity']}\n")
                    f.write(f"Type:           {vuln['Type']}\n")
                    f.write(f"Fixed In:       {vuln.get('FixedInVersion', 'N/A')}\n")
                    f.write("-" * 80 + "\n\n")
            
            # All Scanned Images
            f.write("ALL SCANNED IMAGES\n")
            f.write("=" * 80 + "\n")
            for image in sorted(self.scanned_images):
                count = all_images_with_counts.get(image, 0)
                platform = self.get_image_platform(image)
                platform_info = f" [{platform}]" if platform else ""
                status = "✅" if count == 0 else "⚠️"
                f.write(f"{status} {image}{platform_info}\n")
                f.write(f"   └─ {count} vulnerabilities\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("END OF REPORT\n".center(80))
            f.write("=" * 80 + "\n")
        
        return filepath
    
    def save_html_reports(self):
        """Save HTML reports for both categories."""
        print("\n🌐 Saving HTML reports...")
        
        # Check which image types were actually scanned
        has_chainguard_images = any(self.is_chainguard_image(img) for img in self.scanned_images)
        has_legacy_images = any(not self.is_chainguard_image(img) for img in self.scanned_images)
        
        # Only report on categories that were actually scanned
        if has_chainguard_images:
            filepath = self.write_html_report(self.chainguard_results, 'grype-chainguard-images.html', 'Chainguard')
            print(f"   ✅ Chainguard HTML: {filepath}")
        
        if has_legacy_images:
            filepath = self.write_html_report(self.legacy_results, 'grype-legacy-images.html', 'Legacy')
            print(f"   ✅ Legacy HTML: {filepath}")
    
    def save_text_reports(self):
        """Save text reports for both categories."""
        print("\n📄 Saving text reports...")
        
        # Check which image types were actually scanned
        has_chainguard_images = any(self.is_chainguard_image(img) for img in self.scanned_images)
        has_legacy_images = any(not self.is_chainguard_image(img) for img in self.scanned_images)
        
        # Only report on categories that were actually scanned
        if has_chainguard_images:
            filepath = self.write_text_report(self.chainguard_results, 'grype-chainguard-images.txt', 'Chainguard')
            print(f"   ✅ Chainguard text: {filepath}")
        
        if has_legacy_images:
            filepath = self.write_text_report(self.legacy_results, 'grype-legacy-images.txt', 'Legacy')
            print(f"   ✅ Legacy text: {filepath}")
    
    def generate_excel_report(self, csv_file: Path):
        """Generate Excel report from CSV file."""
        if not EXCEL_AVAILABLE:
            return
        
        print(f"\n📊 Generating Excel report for {csv_file.name}...")
        
        # Load CSV data
        data = []
        with open(csv_file, 'r') as f:
            import csv
            reader = csv.DictReader(f)
            for row in reader:
                data.append({k: v.strip('"') for k, v in row.items()})
        
        if not data:
            print("   ⚠️  No data to generate report")
            return
        
        # Determine report type from filename
        report_type = "Chainguard" if "chainguard" in csv_file.name.lower() else "Legacy"
        
        # Create workbook
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create summary sheet
        self._create_summary_sheet(wb, data, report_type)
        
        # Create severity sheets (always create all, even if empty)
        for severity in self.SEVERITY_ORDER:
            filtered = [d for d in data if d['Severity'] == severity]
            self._create_details_sheet(wb, filtered, severity)
        
        # Create full details
        self._create_details_sheet(wb, data, "Full Details")
        
        # Create per-image sheets
        self._create_image_sheets(wb, data)
        
        # Save workbook
        output_file = csv_file.with_suffix('.xlsx')
        wb.save(output_file)
        print(f"   ✅ Excel report saved: {output_file}")
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict], report_type: str = ""):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        stats = {
            'total': len(data),
            'by_severity': Counter(d['Severity'] for d in data),
            'by_image': Counter(d['Image'] for d in data),
            'by_type': Counter(d['Type'] for d in data),
            'unique_cves': len(set(d['Vulnerability'] for d in data)),
            'unique_packages': len(set(d['Package'] for d in data)),
            'unique_images': len(set(d['Image'] for d in data)),
            'fixable': sum(1 for d in data if d.get('FixedInVersion', 'N/A') != 'N/A'),
        }
        
        # Title with report type
        title = f"{report_type} Security Vulnerability Scan Report" if report_type else "Security Vulnerability Scan Report"
        ws['A1'] = title
        ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {self.scan_timestamp}"
        ws['A2'].font = Font(italic=True, color='666666')
        ws.merge_cells('A2:D2')
        
        # Key metrics
        row = 4
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        metrics = [
            ("Total Vulnerabilities:", stats['total']),
            ("Unique CVEs:", stats['unique_cves']),
            ("Affected Packages:", stats['unique_packages']),
            ("Affected Images:", stats['unique_images']),
            ("Fixable Issues:", f"{stats['fixable']} ({stats['fixable']/stats['total']*100:.1f}%)" if stats['total'] > 0 else "0 (0%)"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Severity breakdown
        row += 1
        ws[f'A{row}'] = "BY SEVERITY"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        severity_start_row = row
        severity_chart_start = None  # Track first non-zero for chart
        severity_chart_end = None    # Track last non-zero for chart
        
        for severity in self.SEVERITY_ORDER:
            count = stats['by_severity'].get(severity, 0)
            # Always show Critical, High, Medium, Low even if 0
            if count > 0 or severity in ['Critical', 'High', 'Medium', 'Low']:
                ws[f'A{row}'] = severity
                ws[f'B{row}'] = count
                
                # Track rows with actual data for chart
                if count > 0:
                    if severity_chart_start is None:
                        severity_chart_start = row
                    severity_chart_end = row
                
                # Color code
                color = self.SEVERITY_COLORS.get(severity, 'FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[f'A{row}'].font = Font(bold=True, color='FFFFFF' if severity in ['Critical', 'High'] else '000000')
                
                row += 1
        
        severity_end_row = row - 1
        
        # Image breakdown (show ALL scanned images, including those with 0 vulns)
        row += 1
        ws[f'A{row}'] = "BY IMAGE"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Get all images (including those with 0 vulnerabilities)
        all_images_with_counts = {}
        for image in self.scanned_images:
            all_images_with_counts[image] = stats['by_image'].get(image, 0)
        
        # Sort by count (descending), then by name
        for image, count in sorted(all_images_with_counts.items(), key=lambda x: (-x[1], x[0])):
            platform = self.get_image_platform(image)
            platform_info = f" [{platform}]" if platform else ""
            ws[f'A{row}'] = f"{image}{platform_info}"
            ws[f'B{row}'] = count
            row += 1
        
        # Package type breakdown
        row += 1
        ws[f'A{row}'] = "BY PACKAGE TYPE"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        for pkg_type, count in sorted(stats['by_type'].items(), key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = pkg_type
            ws[f'B{row}'] = count
            row += 1
        
        # Add pie chart for severity (without emoji in title)
        if stats['total'] > 0 and severity_chart_start is not None and severity_chart_end is not None:
            # Add a title above the chart in the spreadsheet (no emoji)
            ws['D3'] = "Vulnerabilities by Severity"
            ws['D3'].font = Font(size=12, bold=True, color='1F4E78')
            ws.merge_cells('D3:G3')
            ws['D3'].alignment = Alignment(horizontal='center')
            
            try:
                chart = PieChart()
                chart.title = None  # No title on chart itself
                chart.height = 10
                chart.width = 15
                
                # Reference only rows with non-zero data
                labels = Reference(ws, min_col=1, min_row=severity_chart_start, max_row=severity_chart_end)
                data = Reference(ws, min_col=2, min_row=severity_chart_start, max_row=severity_chart_end)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(labels)
                
                # Note: Custom colors per severity would require complex openpyxl API
                # Excel will auto-assign colors based on data order
                # Colors may differ between Legacy and Chainguard charts
                
                ws.add_chart(chart, "D4")
            except Exception as e:
                # If chart fails, just skip it
                ws['D4'] = f"Chart unavailable"
                ws['D4'].font = Font(italic=True, color='666666')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_details_sheet(self, wb: Workbook, data: List[Dict], name: str):
        """Create detailed vulnerability sheet."""
        sheet_name = f"{name}"[:31]  # Remove emoji
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Image', 'Platform', 'Package', 'Version', 'Vulnerability', 'Severity', 'Type', 'Fixed In']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Check if data is empty
        if not data:
            # Add "No vulnerabilities" message
            ws['A2'] = f"✅ No {name} vulnerabilities found!"
            ws['A2'].font = Font(size=14, bold=True, color='228B22')
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            widths = [35, 15, 30, 15, 20, 12, 10, 15]
            for col_num, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            return
        
        # Sort by severity
        severity_map = {sev: idx for idx, sev in enumerate(self.SEVERITY_ORDER)}
        sorted_data = sorted(data, key=lambda x: (severity_map.get(x['Severity'], 999), x['Image']))
        
        # Data rows
        for row_num, row_data in enumerate(sorted_data, 2):
            ws.cell(row=row_num, column=1, value=row_data['Image'])
            ws.cell(row=row_num, column=2, value=row_data.get('Platform', 'unknown'))
            ws.cell(row=row_num, column=3, value=row_data['Package'])
            ws.cell(row=row_num, column=4, value=row_data['Version'])
            
            # Hyperlinked CVE
            cve = row_data['Vulnerability']
            cve_cell = ws.cell(row=row_num, column=5, value=cve)
            if cve.startswith('CVE-'):
                cve_cell.hyperlink = f"https://nvd.nist.gov/vuln/detail/{cve}"
                cve_cell.font = Font(color='0563C1', underline='single')
            elif cve.startswith('GHSA-'):
                cve_cell.hyperlink = f"https://github.com/advisories/{cve}"
                cve_cell.font = Font(color='0563C1', underline='single')
            
            severity = row_data['Severity']
            severity_cell = ws.cell(row=row_num, column=6, value=severity)
            color = self.SEVERITY_COLORS.get(severity, 'FFFFFF')
            severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            if severity in ['Critical', 'High']:
                severity_cell.font = Font(bold=True, color='FFFFFF')
            
            ws.cell(row=row_num, column=7, value=row_data['Type'])
            ws.cell(row=row_num, column=8, value=row_data.get('FixedInVersion', 'N/A'))
        
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = ws['A2']
        
        # Column widths
        widths = [35, 15, 30, 15, 20, 12, 10, 15]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    def _create_image_sheets(self, wb: Workbook, data: List[Dict]):
        """Create per-image vulnerability sheets for ALL scanned images."""
        # Use scanned_images to include images with 0 vulnerabilities
        for image in sorted(self.scanned_images):
            # Shorten image name for sheet title (max 31 chars)
            sheet_name = image.replace(':', '_').replace('/', '_')
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            # Remove emoji for Excel compatibility
            
            image_data = [d for d in data if d['Image'] == image]
            
            ws = wb.create_sheet(sheet_name)
            
            # Get platform for this image
            platform = self.get_image_platform(image)
            platform_info = f" [{platform}]" if platform else ""
            
            # Title
            ws['A1'] = f"Image: {image}{platform_info}"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')
            
            # Check if image has any vulnerabilities
            if not image_data:
                # No vulnerabilities found for this image
                ws['A2'] = "Total: 0 vulnerabilities"
                ws['A2'].font = Font(italic=True)
                ws.merge_cells('A2:F2')
                
                # Headers still
                headers = ['Package', 'Version', 'Vulnerability', 'Severity', 'Type', 'Fixed In']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_num, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Add success message
                ws['A5'] = "No vulnerabilities detected in this image!"
                ws['A5'].font = Font(size=12, bold=True, color='228B22')
                ws.merge_cells('A5:F5')
                ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Stats
                stats = Counter(d['Severity'] for d in image_data)
                stat_text = f"Total: {len(image_data)} | "
                for sev in self.SEVERITY_ORDER:
                    if stats[sev] > 0:
                        stat_text += f"{sev}: {stats[sev]} | "
                ws['A2'] = stat_text.rstrip('| ')
                ws['A2'].font = Font(italic=True)
                ws.merge_cells('A2:F2')
                
                # Headers
                headers = ['Package', 'Version', 'Vulnerability', 'Severity', 'Type', 'Fixed In']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_num, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Sort by severity
                severity_map = {sev: idx for idx, sev in enumerate(self.SEVERITY_ORDER)}
                sorted_data = sorted(
                    image_data,
                    key=lambda x: (severity_map.get(x['Severity'], 999), x['Package'])
                )
                
                # Data rows
                for row_num, row_data in enumerate(sorted_data, 5):
                    ws.cell(row=row_num, column=1, value=row_data['Package'])
                    ws.cell(row=row_num, column=2, value=row_data['Version'])
                    
                    cve = row_data['Vulnerability']
                    cve_cell = ws.cell(row=row_num, column=3, value=cve)
                    if cve.startswith('CVE-'):
                        cve_cell.hyperlink = f"https://nvd.nist.gov/vuln/detail/{cve}"
                        cve_cell.font = Font(color='0563C1', underline='single')
                    elif cve.startswith('GHSA-'):
                        cve_cell.hyperlink = f"https://github.com/advisories/{cve}"
                        cve_cell.font = Font(color='0563C1', underline='single')
                    
                    severity = row_data['Severity']
                    severity_cell = ws.cell(row=row_num, column=4, value=severity)
                    color = self.SEVERITY_COLORS.get(severity, 'FFFFFF')
                    severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    if severity in ['Critical', 'High']:
                        severity_cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        severity_cell.font = Font(bold=True)
                    
                    ws.cell(row=row_num, column=5, value=row_data['Type'])
                    ws.cell(row=row_num, column=6, value=row_data.get('FixedInVersion', 'N/A'))
            
            # Freeze and filter
            ws.freeze_panes = ws['A5']
            ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
            
            # Column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
    
    def generate_comparison_report(self):
        """Generate comparison report if both CSV files exist."""
        if not EXCEL_AVAILABLE:
            return
        
        legacy_csv = self.output_dir / 'grype-legacy-images.csv'
        cg_csv = self.output_dir / 'grype-chainguard-images.csv'
        
        if not legacy_csv.exists() or not cg_csv.exists():
            return
        
        print("\n📊 Generating comparison report...")
        
        # Load both datasets
        legacy_data = self._load_csv(legacy_csv)
        cg_data = self._load_csv(cg_csv)
        
        # Create comparison workbook
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create executive summary
        self._create_comparison_summary(wb, legacy_data, cg_data)
        
        # Save
        output_file = self.output_dir / f"comparison-report-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_file)
        print(f"   ✅ Comparison report saved: {output_file}")
    
    def generate_html_comparison_report(self):
        """Generate HTML comparison report if both result sets exist."""
        if not self.legacy_results or not self.chainguard_results:
            return
        
        print("\n🌐 Generating HTML comparison report...")
        
        # Calculate stats
        legacy_stats = {
            'total': len(self.legacy_results),
            'by_severity': Counter(d['Severity'] for d in self.legacy_results),
            'unique_cves': len(set(d['Vulnerability'] for d in self.legacy_results)),
        }
        
        cg_stats = {
            'total': len(self.chainguard_results),
            'by_severity': Counter(d['Severity'] for d in self.chainguard_results),
            'unique_cves': len(set(d['Vulnerability'] for d in self.chainguard_results)),
        }
        
        filepath = self.output_dir / f"comparison-report-{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Security Comparison: Legacy vs Chainguard</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f7fa;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        
        h1 {{
            color: #1F4E78;
            margin-bottom: 10px;
            font-size: 2.5em;
            text-align: center;
        }}
        
        .timestamp {{
            color: #666;
            font-style: italic;
            margin-bottom: 40px;
            text-align: center;
        }}
        
        .comparison-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin-bottom: 40px;
        }}
        
        .comparison-card {{
            padding: 30px;
            border-radius: 8px;
            color: white;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        .legacy-card {{
            background: linear-gradient(135deg, #DC143C 0%, #8B0000 100%);
        }}
        
        .chainguard-card {{
            background: linear-gradient(135deg, #28a745 0%, #155724 100%);
        }}
        
        .comparison-card h2 {{
            color: white;
            margin-bottom: 20px;
            font-size: 1.8em;
            border: none;
        }}
        
        .stat-row {{
            display: flex;
            justify-content: space-between;
            padding: 10px 0;
            border-bottom: 1px solid rgba(255,255,255,0.2);
        }}
        
        .stat-row:last-child {{
            border-bottom: none;
        }}
        
        .stat-label {{
            font-weight: normal;
            opacity: 0.9;
        }}
        
        .stat-value {{
            font-weight: bold;
            font-size: 1.3em;
        }}
        
        h2 {{
            color: #1F4E78;
            margin: 40px 0 20px 0;
            padding-bottom: 10px;
            border-bottom: 2px solid #e0e0e0;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        
        thead {{
            background: #1F4E78;
            color: white;
        }}
        
        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        
        td {{
            padding: 15px;
            border-bottom: 1px solid #e0e0e0;
            text-align: center;
        }}
        
        td:first-child {{
            text-align: left;
            font-weight: 600;
        }}
        
        tbody tr:hover {{
            background: #f8f9fa;
        }}
        
        .improvement-positive {{
            background: #90EE90;
            padding: 5px 10px;
            border-radius: 4px;
            font-weight: 600;
        }}
        
        .improvement-moderate {{
            background: #FFD700;
            padding: 5px 10px;
            border-radius: 4px;
            font-weight: 600;
        }}
        
        .improvement-none {{
            color: #999;
        }}
        
        .reduction {{
            font-weight: 600;
            color: #28a745;
        }}
        
        .summary-box {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 8px;
            text-align: center;
            margin: 40px 0;
        }}
        
        .summary-box h2 {{
            color: white;
            border: none;
            font-size: 2em;
            margin-bottom: 15px;
        }}
        
        .summary-box p {{
            font-size: 1.3em;
            margin: 10px 0;
        }}
        
        .highlight {{
            font-size: 2.5em;
            font-weight: bold;
            display: block;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔒 Security Comparison Report</h1>
        <h2 style="text-align: center; color: #666; font-size: 1.2em; border: none; margin-top: 5px;">Legacy Images vs Chainguard Images</h2>
        <div class="timestamp">Generated: {self.scan_timestamp}</div>
        
        <div class="comparison-grid">
            <div class="comparison-card legacy-card">
                <h2>📦 Legacy Images</h2>
                <div class="stat-row">
                    <span class="stat-label">Total Vulnerabilities</span>
                    <span class="stat-value">{legacy_stats['total']}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Unique CVEs</span>
                    <span class="stat-value">{legacy_stats['unique_cves']}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Critical</span>
                    <span class="stat-value">{legacy_stats['by_severity'].get('Critical', 0)}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">High</span>
                    <span class="stat-value">{legacy_stats['by_severity'].get('High', 0)}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Medium</span>
                    <span class="stat-value">{legacy_stats['by_severity'].get('Medium', 0)}</span>
                </div>
            </div>
            
            <div class="comparison-card chainguard-card">
                <h2>🛡️ Chainguard Images</h2>
                <div class="stat-row">
                    <span class="stat-label">Total Vulnerabilities</span>
                    <span class="stat-value">{cg_stats['total']}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Unique CVEs</span>
                    <span class="stat-value">{cg_stats['unique_cves']}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Critical</span>
                    <span class="stat-value">{cg_stats['by_severity'].get('Critical', 0)}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">High</span>
                    <span class="stat-value">{cg_stats['by_severity'].get('High', 0)}</span>
                </div>
                <div class="stat-row">
                    <span class="stat-label">Medium</span>
                    <span class="stat-value">{cg_stats['by_severity'].get('Medium', 0)}</span>
                </div>
            </div>
        </div>
"""
        
        # Calculate overall improvement
        if legacy_stats['total'] > 0:
            total_reduction = legacy_stats['total'] - cg_stats['total']
            total_improvement = (total_reduction / legacy_stats['total']) * 100
            
            html += f"""
        <div class="summary-box">
            <h2>✨ Security Improvement</h2>
            <p>By switching to Chainguard images, you reduced vulnerabilities by:</p>
            <span class="highlight">{total_improvement:.1f}%</span>
            <p>That's <strong>{total_reduction}</strong> fewer vulnerabilities!</p>
        </div>
"""
        
        html += """
        <h2>Detailed Comparison</h2>
        <table>
            <thead>
                <tr>
                    <th>Metric</th>
                    <th>Legacy</th>
                    <th>Chainguard</th>
                    <th>Reduction</th>
                    <th>% Improvement</th>
                </tr>
            </thead>
            <tbody>
"""
        
        metrics = [
            ("Total Vulnerabilities", legacy_stats['total'], cg_stats['total']),
            ("Unique CVEs", legacy_stats['unique_cves'], cg_stats['unique_cves']),
            ("Critical Severity", legacy_stats['by_severity'].get('Critical', 0), cg_stats['by_severity'].get('Critical', 0)),
            ("High Severity", legacy_stats['by_severity'].get('High', 0), cg_stats['by_severity'].get('High', 0)),
            ("Medium Severity", legacy_stats['by_severity'].get('Medium', 0), cg_stats['by_severity'].get('Medium', 0)),
        ]
        
        for label, legacy_val, cg_val in metrics:
            reduction = legacy_val - cg_val
            
            if legacy_val > 0:
                improvement = (reduction / legacy_val) * 100
                improvement_class = "improvement-positive" if improvement >= 90 else "improvement-moderate" if improvement >= 50 else ""
                improvement_text = f'<span class="{improvement_class}">{improvement:.1f}%</span>'
            else:
                improvement_text = '<span class="improvement-none">N/A</span>'
            
            html += f"""
                <tr>
                    <td>{label}</td>
                    <td>{legacy_val}</td>
                    <td>{cg_val}</td>
                    <td class="reduction">{reduction}</td>
                    <td>{improvement_text}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
        
        <div style="margin-top: 40px; padding: 20px; background: #f8f9fa; border-radius: 6px; border-left: 4px solid #667eea;">
            <h3 style="color: #1F4E78; margin-bottom: 10px;">📊 About This Report</h3>
            <p style="color: #666;">
                This comparison report analyzes security vulnerabilities found in legacy container images
                versus Chainguard's hardened container images. The data is generated using Grype security scanner
                and represents vulnerabilities detected at scan time. Regular scanning is recommended to maintain
                security posture.
            </p>
        </div>
    </div>
</body>
</html>
"""
        
        with open(filepath, 'w') as f:
            f.write(html)
        
        print(f"   ✅ HTML comparison report saved: {filepath}")
    
    def generate_text_comparison_report(self):
        """Generate text comparison report if both result sets exist."""
        if not self.legacy_results or not self.chainguard_results:
            return
        
        print("\n📄 Generating text comparison report...")
        
        # Calculate stats
        legacy_stats = {
            'total': len(self.legacy_results),
            'by_severity': Counter(d['Severity'] for d in self.legacy_results),
            'unique_cves': len(set(d['Vulnerability'] for d in self.legacy_results)),
        }
        
        cg_stats = {
            'total': len(self.chainguard_results),
            'by_severity': Counter(d['Severity'] for d in self.chainguard_results),
            'unique_cves': len(set(d['Vulnerability'] for d in self.chainguard_results)),
        }
        
        filepath = self.output_dir / f"comparison-report-{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        with open(filepath, 'w') as f:
            # Header
            f.write("=" * 80 + "\n")
            f.write("SECURITY COMPARISON REPORT\n".center(80))
            f.write("Legacy Images vs Chainguard Images\n".center(80))
            f.write("=" * 80 + "\n")
            f.write(f"Generated: {self.scan_timestamp}\n".center(80))
            f.write("=" * 80 + "\n\n")
            
            # Side-by-side summary
            f.write("SUMMARY STATISTICS\n")
            f.write("-" * 80 + "\n")
            f.write(f"{'Metric':<30} {'Legacy':>15} {'Chainguard':>15} {'Reduction':>15}\n")
            f.write("-" * 80 + "\n")
            
            metrics = [
                ("Total Vulnerabilities", legacy_stats['total'], cg_stats['total']),
                ("Unique CVEs", legacy_stats['unique_cves'], cg_stats['unique_cves']),
                ("Critical Severity", legacy_stats['by_severity'].get('Critical', 0), cg_stats['by_severity'].get('Critical', 0)),
                ("High Severity", legacy_stats['by_severity'].get('High', 0), cg_stats['by_severity'].get('High', 0)),
                ("Medium Severity", legacy_stats['by_severity'].get('Medium', 0), cg_stats['by_severity'].get('Medium', 0)),
                ("Low Severity", legacy_stats['by_severity'].get('Low', 0), cg_stats['by_severity'].get('Low', 0)),
            ]
            
            for label, legacy_val, cg_val in metrics:
                reduction = legacy_val - cg_val
                f.write(f"{label:<30} {legacy_val:>15} {cg_val:>15} {reduction:>15}\n")
            
            f.write("\n")
            
            # Improvement analysis
            f.write("IMPROVEMENT ANALYSIS\n")
            f.write("-" * 80 + "\n")
            
            for label, legacy_val, cg_val in metrics:
                reduction = legacy_val - cg_val
                if legacy_val > 0:
                    improvement = (reduction / legacy_val) * 100
                    icon = "✅" if improvement >= 90 else "⭐" if improvement >= 50 else "📊"
                    f.write(f"{icon} {label:<30}: {improvement:>6.1f}% improvement\n")
                else:
                    f.write(f"  {label:<30}: N/A (no vulnerabilities)\n")
            
            f.write("\n")
            
            # Overall summary
            if legacy_stats['total'] > 0:
                total_reduction = legacy_stats['total'] - cg_stats['total']
                total_improvement = (total_reduction / legacy_stats['total']) * 100
                
                f.write("OVERALL SECURITY IMPROVEMENT\n")
                f.write("=" * 80 + "\n")
                f.write(f"\n")
                f.write(f"  By switching to Chainguard images, you reduced vulnerabilities by:\n")
                f.write(f"\n")
                f.write(f"    🎯 {total_improvement:.1f}% reduction in total vulnerabilities\n")
                f.write(f"    🛡️  {total_reduction} fewer vulnerabilities to manage\n")
                f.write(f"\n")
                
                # Critical and High analysis
                crit_reduction = legacy_stats['by_severity'].get('Critical', 0) - cg_stats['by_severity'].get('Critical', 0)
                high_reduction = legacy_stats['by_severity'].get('High', 0) - cg_stats['by_severity'].get('High', 0)
                
                if crit_reduction > 0 or high_reduction > 0:
                    f.write(f"  Most Significant Improvements:\n")
                    if crit_reduction > 0:
                        f.write(f"    🔴 Critical: {crit_reduction} fewer critical vulnerabilities\n")
                    if high_reduction > 0:
                        f.write(f"    🟠 High: {high_reduction} fewer high-severity vulnerabilities\n")
                f.write(f"\n")
            
            f.write("=" * 80 + "\n")
            f.write("\n")
            f.write("ABOUT THIS REPORT\n")
            f.write("-" * 80 + "\n")
            f.write("This comparison report analyzes security vulnerabilities found in legacy\n")
            f.write("container images versus Chainguard's hardened container images.\n")
            f.write("\n")
            f.write("The data is generated using Grype security scanner and represents\n")
            f.write("vulnerabilities detected at scan time. Regular scanning is recommended\n")
            f.write("to maintain security posture.\n")
            f.write("\n")
            f.write("=" * 80 + "\n")
            f.write("END OF COMPARISON REPORT\n".center(80))
            f.write("=" * 80 + "\n")
        
        print(f"   ✅ Text comparison report saved: {filepath}")
    
    def _load_csv(self, filepath: Path) -> List[Dict]:
        """Load CSV file into list of dicts."""
        data = []
        with open(filepath, 'r') as f:
            import csv
            reader = csv.DictReader(f)
            for row in reader:
                data.append({k: v.strip('"') for k, v in row.items()})
        return data
    
    def _create_comparison_summary(self, wb: Workbook, legacy_data: List[Dict], cg_data: List[Dict]):
        """Create comparison summary sheet."""
        ws = wb.create_sheet("Comparison", 0)
        
        # Calculate stats
        legacy_stats = {
            'total': len(legacy_data),
            'by_severity': Counter(d['Severity'] for d in legacy_data),
            'unique_cves': len(set(d['Vulnerability'] for d in legacy_data)),
        }
        
        cg_stats = {
            'total': len(cg_data),
            'by_severity': Counter(d['Severity'] for d in cg_data),
            'unique_cves': len(set(d['Vulnerability'] for d in cg_data)),
        }
        
        # Title
        ws['A1'] = "Security Comparison: Legacy vs Chainguard"
        ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Generated: {self.scan_timestamp}"
        ws['A2'].font = Font(italic=True, color='666666')
        ws.merge_cells('A2:E2')
        
        # Headers
        row = 4
        ws[f'B{row}'] = "Legacy"
        ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'B{row}'].fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'C{row}'] = "Chainguard"
        ws[f'C{row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'C{row}'].fill = PatternFill(start_color='228B22', end_color='228B22', fill_type='solid')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = "Reduction"
        ws[f'D{row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'D{row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'E{row}'] = "% Improvement"
        ws[f'E{row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'E{row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        
        # Metrics
        row += 1
        metrics = [
            ("Total Vulnerabilities", legacy_stats['total'], cg_stats['total']),
            ("Unique CVEs", legacy_stats['unique_cves'], cg_stats['unique_cves']),
            ("Critical", legacy_stats['by_severity'].get('Critical', 0), cg_stats['by_severity'].get('Critical', 0)),
            ("High", legacy_stats['by_severity'].get('High', 0), cg_stats['by_severity'].get('High', 0)),
            ("Medium", legacy_stats['by_severity'].get('Medium', 0), cg_stats['by_severity'].get('Medium', 0)),
        ]
        
        for label, legacy_val, cg_val in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = legacy_val
            ws[f'C{row}'] = cg_val
            
            reduction = legacy_val - cg_val
            ws[f'D{row}'] = reduction
            
            if legacy_val > 0:
                improvement = (reduction / legacy_val) * 100
                ws[f'E{row}'] = f"{improvement:.1f}%"
                
                if improvement >= 90:
                    ws[f'E{row}'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif improvement >= 50:
                    ws[f'E{row}'].fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            else:
                ws[f'E{row}'] = "N/A"
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
    
    def generate_reports(self):
        """Generate all reports in multiple formats."""
        # Always generate CSV, HTML, and text reports
        self.save_csv_reports()
        self.save_html_reports()
        self.save_text_reports()
        
        # Generate comparison reports if both sets exist
        self.generate_html_comparison_report()
        self.generate_text_comparison_report()
        
        # Generate Excel reports if openpyxl is available
        if not EXCEL_AVAILABLE:
            print("\n💡 TIP: Install openpyxl to generate Excel reports:")
            print("   pip install -r ./scanners/requirements.txt")
            return
        
        # Generate individual Excel reports
        if self.chainguard_results:
            csv_file = self.output_dir / 'grype-chainguard-images.csv'
            if csv_file.exists():
                self.generate_excel_report(csv_file)
        
        if self.legacy_results:
            csv_file = self.output_dir / 'grype-legacy-images.csv'
            if csv_file.exists():
                self.generate_excel_report(csv_file)
        
        # Generate Excel comparison report
        self.generate_comparison_report()
    
    def run(self):
        """Main execution flow."""
        print("🚀 Container Security Scanner")
        print("=" * 60)
        
        # Check dependencies
        if not self.check_dependencies():
            return 1
        
        # Scan containers
        self.scan_all_containers()
        
        # Check if we found anything
        if not self.chainguard_results and not self.legacy_results:
            print("\n✅ No vulnerabilities found in any running containers! 🎉")
            return 0
        
        # Generate reports
        self.generate_reports()
        
        print("\n✅ Scan complete!")
        print("=" * 60)
        print("\n📁 Report Formats Generated:")
        print("   • CSV files - for data processing")
        print("   • HTML files - for web viewing (Instruqt-compatible)")
        print("   • Text files - for terminal viewing")
        if EXCEL_AVAILABLE:
            print("   • Excel files - for detailed analysis with charts")
        print("\n💡 Note: Multi-architecture images (e.g., python:latest) may have")
        print("   different vulnerabilities on different platforms (arm64 vs amd64).")
        print("   The Platform column in reports shows which architecture was scanned.")
        print("=" * 60)
        return 0


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description='Scan Docker containers for vulnerabilities and generate reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 scan-and-report.py
  python3 scan-and-report.py --output-dir ./reports
        """
    )
    
    parser.add_argument(
        '--output-dir',
        default='./scanners/scan-results',
        help='Output directory for reports (default: ./scanners/scan-results)'
    )
    
    args = parser.parse_args()
    
    scanner = VulnerabilityScanner(output_dir=args.output_dir)
    sys.exit(scanner.run())


if __name__ == '__main__':
    main()

